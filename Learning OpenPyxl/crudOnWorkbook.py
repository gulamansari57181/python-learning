from openpyxl import Workbook

# initialze a workbook
wb = Workbook()

# create a sheet in a workbbok  with name and index

ws1 = wb.create_sheet("A sheet",0)
# appending a new sheet to an end of existing sheets
ws2 = wb.create_sheet("B sheet")

for sheet in wb:
    print(sheet.title)

# Output : 
# A sheet
# Sheet    //default sheet
# B sheet

# Remove a sheet by name
wb.remove(wb["Sheet"])

print("After removal of sheet by name : ")

# printing sheet name
for sheet in wb:
    print(sheet.title)
    
# Changing sheet title
ws1.title ="First Sheet"
ws2.title = "Second Sheet"

print("After changing the sheet name :")
# printing sheet name
for sheet in wb:
    print(sheet.title)
    
# saving a workbook by a name on a current directory
wb.save("CRUD on Workbook.xlsx")
    